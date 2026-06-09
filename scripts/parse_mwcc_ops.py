"""
parse_mwcc_ops.py — Parses OWNA weekly exports for My World Childcare.

Reads from mwcc-inbox/:
  MYWORLD_REPORT*.xlsx  — revenue, wage %, enquiries, exits, enrolments per centre
  utilisation*.xlsx     — per-room daily occupancy (Mon–Fri) per centre

Saves to state/mwcc-ops.json

Usage:
  python scripts/parse_mwcc_ops.py              # reads from mwcc-inbox/
  python scripts/parse_mwcc_ops.py /custom/path # reads from a custom folder

Drop zone workflow:
  Every Monday before 2pm, drop MYWORLD_REPORT.xlsx + utilisation.xlsx into mwcc-inbox/
  The 2pm cron runs this script automatically.

Data notes:
  - Reporting convention (agreed 07 Jun 2026):
    · OWNA centre operations data → Mon–Fri (centres operate weekdays only).
      For 07 Jun report: 01–05 Jun 2026.
    · Digital marketing (GA4, GSC, Meta, Google Ads, Metricool/social) → Sat–Fri.
      Captures weekend digital activity that centres don't have.
      For 07 Jun report: 30 May – 05 Jun 2026.
  - The parser overrides the OWNA file's native period with the computed Mon-Fri
    window via compute_owna_week(). The original file-reported period is kept
    in output['period']['owna_file_period'] for reference.
  - Wage breach threshold: 42% (wage exc. leave) — NOT surfaced in marketing UI
    (HR/finance concern, not marketing). Available in data for ops tooling.
  - Compliance risk: room occupancy > 100%
  - Centre name mapping: OWNA uses 'WAKIKI' — canonical name is 'Waikiki'
"""

import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional, Tuple


def compute_owna_week():
    """Compute the Mon→Fri window for the last completed working week.

    Convention agreed 07 Jun 2026 (Tia):
      OWNA centre operations data is Mon–Fri because centres only operate
      weekdays. Sat/Sun = closed = always zero. Reporting on Mon–Fri removes
      noise and aligns with how the team actually runs the business.

    Marketing scripts (GA4/GSC/Meta/Ads) report Sat–Fri (the wider 7-day
    window that captures weekend digital activity). These are documented as
    different windows by design.

    Returns (start_date_str, end_date_str) in YYYY-MM-DD format.
    end = last completed Friday
    start = same week's Monday = end - 4 days
    """
    today = datetime.today()
    days_since_friday = (today.weekday() - 4) % 7
    end_date   = today - timedelta(days=days_since_friday)
    start_date = end_date - timedelta(days=4)
    return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")

try:
    import openpyxl
except ImportError:
    print("[MWCC Ops] ERROR: openpyxl not installed.")
    print("           Run: pip install openpyxl")
    sys.exit(1)

BASE_DIR    = Path(__file__).resolve().parent.parent
INBOX_DIR   = BASE_DIR / "mwcc-inbox"
STATE_DIR   = BASE_DIR / "state"
OUTPUT_FILE = STATE_DIR / "mwcc-ops.json"

WAGE_BREACH_THRESHOLD = 42.0  # wage exc. leave % above this = breach

# OWNA centre name → canonical display name
CENTRE_MAP = {
    "ARMADALE OSHC":   "Armadale",
    "MIDVALE":         "Midvale",
    "ROCKINGHAM OSHC": "Rockingham",
    "SEVILLE GROVE":   "Seville Grove",
    "WAKIKI":          "Waikiki",   # OWNA spells it WAKIKI
    "WAIKIKI":         "Waikiki",
}

# Utilisation sheet name → canonical centre name
SHEET_CENTRE_MAP = {
    "Armadale":     "Armadale",
    "Midvale":      "Midvale",
    "Rockingham":   "Rockingham",
    "Seville Grove":"Seville Grove",
    "Wakiki":       "Waikiki",
    "Waikiki":      "Waikiki",
}

# Room name normalisation (lowercase → canonical)
ROOM_MAP = {
    "babies room":   "Babies",
    "toddlers room": "Toddlers",
    "kindy room":    "Kindy",
    "before school": "Before School",
    "after school":  "After School",
    "vacation care": "Vacation",
    "vacation":      "Vacation",
}


# ─────────────────────────────────────────────────────────────────
# File discovery
# ─────────────────────────────────────────────────────────────────

def _find_file(inbox: Path, prefix: str) -> Optional[Path]:
    """Find the most recently modified xlsx matching prefix* in inbox."""
    matches = sorted(inbox.glob(f"{prefix}*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if matches:
        return matches[0]
    exact = inbox / f"{prefix}.xlsx"
    return exact if exact.exists() else None


# ─────────────────────────────────────────────────────────────────
# MYWORLD_REPORT parser
# ─────────────────────────────────────────────────────────────────

def _parse_myworld_report(path: Path) -> Tuple[Optional[dict], Optional[str], Optional[str]]:
    """
    Parse MYWORLD_REPORT.xlsx 'Weekly Report' sheet.

    Columns (confirmed from live file):
      1  Start Date             — the reporting week start (Monday)
      2  End Date               — the reporting week end (Sunday)
      3  Center                 — OWNA centre name
      4  Wage Inc Leave (%)
      5  Wage Exc Leave (%)     ← primary wage health metric
      6  Revenue ($)
      7  Roster Cost ($)
      8  Leave ($)
      9  Babies (%)             ← room occupancy %
     10  Todds (%)
     11  Kindy (%)
     12  B/S (%)                ← Before School
     13  A/S (%)                ← After School
     14  Vacation (%)
     15  Overall (%)            ← blended centre occupancy
     28  Enquiries
     29  Exits
     30  Enrollments

    Strategy: find max(Start Date), extract all 5 centre rows for that date.
    """
    print(f"[MWCC Ops] Parsing MYWORLD_REPORT: {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Weekly Report" not in wb.sheetnames:
        print(f"[MWCC Ops]   ⚠️  Sheet 'Weekly Report' not found. Sheets: {wb.sheetnames}")
        return None, None, None

    ws = wb["Weekly Report"]

    # --- Find most recent Start Date ---
    max_date = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            d = row[0].date() if hasattr(row[0], "date") else datetime.fromisoformat(str(row[0])).date()
            if max_date is None or d > max_date:
                max_date = d
        except Exception:
            continue

    if not max_date:
        print("[MWCC Ops]   ⚠️  No valid dates found in MYWORLD_REPORT")
        return None, None, None

    print(f"[MWCC Ops]   Reporting week: {max_date}")

    def _f(v):
        try: return float(v or 0)
        except: return 0.0

    def _i(v):
        try: return int(float(v or 0))
        except: return 0

    result      = {}
    period_start = None
    period_end   = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            d = row[0].date() if hasattr(row[0], "date") else datetime.fromisoformat(str(row[0])).date()
        except Exception:
            continue
        if d != max_date:
            continue

        # Set period from first matching row
        if period_start is None:
            period_start = str(row[0])[:10]
            period_end   = str(row[1])[:10]

        centre_raw  = str(row[2]).strip().upper()
        centre_name = CENTRE_MAP.get(centre_raw, centre_raw.title())

        wage_exc = _f(row[4])
        wage_inc = _f(row[3])
        revenue  = _f(row[5])
        roster   = _f(row[6])
        leave    = _f(row[7])
        enrolments = _i(row[29])
        exits      = _i(row[28])

        result[centre_name] = {
            "name":               centre_name,
            "raw_name":           centre_raw,
            "revenue":            round(revenue, 2),
            "wage_inc_leave_pct": round(wage_inc, 2),
            "wage_exc_leave_pct": round(wage_exc, 2),
            "roster_cost":        round(roster, 2),
            "leave_cost":         round(leave, 2),
            "wage_breach":        wage_exc > WAGE_BREACH_THRESHOLD,
            # Room occupancy % from MYWORLD_REPORT (weekly avg, already rounded)
            "occupancy": {
                "Babies":         _i(row[8]),
                "Toddlers":       _i(row[9]),
                "Kindy":          _i(row[10]),
                "Before School":  _i(row[11]),
                "After School":   _i(row[12]),
                "Vacation":       _i(row[13]),
                "Overall":        _i(row[14]),
            },
            "enquiries":    _i(row[27]),
            "exits":        exits,
            "enrolments":   enrolments,
            "net_movement": enrolments - exits,
            # ── Projection data (cols 15-26) for management report ──
            # OWNA's "This Week" + "Next Week" columns are forward-looking
            # projections relative to the report's reference date.
            # Mapped here as next_week_projection + week_after_projection
            # so the management report can show 3-column "actuals → next →
            # week-after" comparison per centre.
            "this_week_projection": {
                "date":      str(row[15])[:10] if row[15] else None,
                "wage_inc_leave_pct":   _f(row[16]),
                "revenue":              _f(row[17]),
                "leave_cost":           _f(row[18]),
                "roster_cost":          _f(row[19]),
                "overall_occupancy":    _i(row[20]),
            },
            "next_week_projection": {
                "date":      str(row[21])[:10] if row[21] else None,
                "wage_inc_leave_pct":   _f(row[22]),
                "revenue":              _f(row[23]),
                "leave_cost":           _f(row[24]),
                "roster_cost":          _f(row[25]),
                "overall_occupancy":    _i(row[26]),
            },
        }

    print(f"[MWCC Ops]   Centres loaded: {list(result.keys())}")
    return result, period_start, period_end


# ─────────────────────────────────────────────────────────────────
# Utilisation parser
# ─────────────────────────────────────────────────────────────────

def _parse_utilisation(path: Path) -> dict:
    """
    Parse utilisation xlsx → per-centre, per-room occupancy detail.

    Structure (one sheet per centre):
      Each sheet has multiple week blocks, each block:
        Header row  : ['Service Name', 'Mon DD, YYYY', 'Daily Capacity', 'MON...', ...]
        Room rows   : [service, room_name, capacity, mon, tue, wed, thu, fri, sat, sun]
        Total row   : ['', '', 'Total', ...]
        Occ % row   : ['', '', 'Daily Occupancy %', '55%', ...]
        Blank row   : (separator between weeks)

    Strategy: take the LAST header block (most recent week).
    Occupancy % = avg(Mon–Fri occupied) / capacity × 100

    Per-room overflow handling (Kelley rule, 09 Jun 2026):
      Per-room occupancy > 100% is NOT a compliance breach. Kelley moves
      children between rooms within ratio (e.g. older Toddlers → Babies'
      room if it has headroom). True licensed-capacity breach is a
      CENTRE-LEVEL check: total bodies under the roof > licensed centre
      capacity. We flag per-room overflow as `staffing_rebalance_needed`
      (operational signal for Kelley), NOT `compliance_risk` (legal).
    """
    print(f"[MWCC Ops] Parsing utilisation: {path.name}")
    wb     = openpyxl.load_workbook(path, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        centre_name = SHEET_CENTRE_MAP.get(sheet_name)
        if not centre_name:
            print(f"[MWCC Ops]   Skipping unknown sheet: '{sheet_name}'")
            continue

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find all header rows (col[0] == 'Service Name')
        header_idx_list = [i for i, r in enumerate(rows) if r[0] == "Service Name"]
        if not header_idx_list:
            print(f"[MWCC Ops]   ⚠️  No week blocks found in sheet '{sheet_name}'")
            continue

        # Take the LAST week block
        last_idx  = header_idx_list[-1]
        header_row = rows[last_idx]
        week_label = str(header_row[1]) if header_row[1] else "unknown"

        # Collect block rows until blank or end of sheet
        block = []
        for row in rows[last_idx + 1:]:
            has_data = any(v is not None and str(v).strip() for v in row)
            if not has_data:
                break
            block.append(row)

        rooms          = {}
        daily_occ_pcts = []

        for row in block:
            col2 = str(row[2] or "").strip()

            # Total row — skip
            if col2 == "Total":
                continue

            # Daily Occupancy % row — capture Mon-Fri trend
            if col2 == "Daily Occupancy %":
                daily_occ_pcts = []
                for v in row[3:8]:   # Mon(3) Tue(4) Wed(5) Thu(6) Fri(7)
                    try:
                        pct = float(str(v or "0").replace("%", "").strip())
                        daily_occ_pcts.append(round(pct, 1))
                    except Exception:
                        daily_occ_pcts.append(0.0)
                continue

            # Room row — col[1] = room name
            room_raw = str(row[1] or "").strip()
            if not room_raw:
                continue

            room_name = ROOM_MAP.get(room_raw.lower(), room_raw)

            try:
                capacity = int(float(row[2] or 0))
            except Exception:
                capacity = 0

            # Mon–Fri values (cols index 3–7); skip Sat/Sun (8–9)
            daily_vals = []
            for v in row[3:8]:
                try:
                    daily_vals.append(int(float(v or 0)))
                except Exception:
                    daily_vals.append(0)

            # Average over active days only (ignore 0s on short weeks)
            active = [v for v in daily_vals if v > 0]
            avg    = round(sum(active) / len(active), 1) if active else 0.0
            occ    = round(avg / capacity * 100, 1) if capacity > 0 else 0.0

            rooms[room_name] = {
                "capacity":                 capacity,
                "avg_daily":                avg,
                "occupancy_pct":            occ,
                "daily_mon_fri":            daily_vals,
                # Per Kelley rule (09 Jun 2026): per-room overflow = staffing
                # rebalance signal, NOT compliance. Kelley shuffles between rooms
                # within ratio. True compliance breach is centre-level (TODO:
                # needs licensed_centre_capacity from Tia per centre).
                "staffing_rebalance_needed": occ > 100.0,
                "compliance_risk":          False,  # CENTRE-LEVEL check; always False at room level
            }

        result[centre_name] = {
            "week_label":     week_label,
            "rooms":          rooms,
            "daily_occ_pcts": daily_occ_pcts,
        }

        room_summary = "  ".join(
            f"{r}: {d['occupancy_pct']}%{'↻' if d.get('staffing_rebalance_needed') else ''}"
            for r, d in rooms.items()
        )
        print(f"[MWCC Ops]   {centre_name}: {room_summary}  (↻ = rebalance signal, NOT compliance)")

    return result


# ─────────────────────────────────────────────────────────────────
# leads-enquiries.xlsx parser — SOURCE OF TRUTH for enquiries (08 Jun 2026)
# ─────────────────────────────────────────────────────────────────
#
# OWNA splits the enquiry pipeline into a dedicated export. Each row is
# one enquiry from a parent with full contact + UTM attribution. We
# extract per-centre counts (both rolling pipeline + this-week new) and
# UTM source breakdown for marketing attribution.

CENTRE_NAME_MAP = {
    "My World Armadale":      "Armadale",
    "My World Midvale":       "Midvale",
    "My World Rockingham":    "Rockingham",
    "My World Seville Grove": "Seville Grove",
    "My World Waikiki":       "Waikiki",
}

def _short_centre(owna_centre_name):
    """Map 'My World Waikiki' → 'Waikiki'. Returns None if unrecognised."""
    if not owna_centre_name:
        return None
    s = str(owna_centre_name).strip()
    return CENTRE_NAME_MAP.get(s)

def _parse_iso_date(s):
    """Parse 'YYYY-MM-DD' or 'May 04, 2026' to datetime.date. Returns None on failure."""
    if not s:
        return None
    s = str(s).strip()
    # Try ISO first (leads-enquiries format)
    try:
        from datetime import datetime
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        pass
    # Try "Mon DD, YYYY" (starters/exits format)
    try:
        from datetime import datetime
        return datetime.strptime(s, "%b %d, %Y").date()
    except ValueError:
        return None

def _parse_leads_enquiries(path, period_start=None, period_end=None):
    """Parse leads-enquiries.xlsx → per-centre enquiry counts.

    Returns dict keyed by short centre name (e.g. 'Waikiki'):
      {
        "Waikiki": {
          "enquiries_this_week": 4,        # filtered to [period_start, period_end]
          "enquiries_pipeline":  11,       # all rows for this centre
          "utm_sources":         {"google": 3, "meta": 1, ...},   # this-week only
          "leads":               [{name, email, phone, submitted, status}, ...],  # this-week only
        },
        ...
      }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[MWCC Ops] ⚠️  openpyxl not installed — cannot parse leads-enquiries.xlsx")
        return {}

    print(f"[MWCC Ops] Parsing leads-enquiries (SOURCE OF TRUTH for enquiries): {path.name}")
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[MWCC Ops] ⚠️  Failed to open leads-enquiries.xlsx: {e}")
        return {}

    if "Data" not in wb.sheetnames:
        print(f"[MWCC Ops] ⚠️  leads-enquiries.xlsx missing 'Data' tab")
        return {}
    ws = wb["Data"]

    from collections import Counter, defaultdict
    result = defaultdict(lambda: {
        "enquiries_this_week": 0,
        "enquiries_pipeline":  0,
        "utm_sources":         Counter(),
        "leads":               [],
    })

    for r in range(2, ws.max_row + 1):
        centre_raw = ws.cell(row=r, column=2).value
        short = _short_centre(centre_raw)
        if not short:
            continue
        result[short]["enquiries_pipeline"] += 1

        submitted_d = _parse_iso_date(ws.cell(row=r, column=15).value)
        in_window = (
            period_start and period_end and submitted_d
            and period_start <= submitted_d <= period_end
        )
        if in_window:
            result[short]["enquiries_this_week"] += 1
            utm_source = (ws.cell(row=r, column=19).value or "").strip() or "(direct)"
            result[short]["utm_sources"][utm_source] += 1
            result[short]["leads"].append({
                "name":      ws.cell(row=r, column=3).value or "",
                "email":     ws.cell(row=r, column=4).value or "",
                "phone":     ws.cell(row=r, column=5).value or "",
                "first_child": ws.cell(row=r, column=6).value or "",
                "submitted": str(submitted_d) if submitted_d else "",
                "status":    ws.cell(row=r, column=18).value or "",
                "utm_source":   utm_source,
                "utm_campaign": (ws.cell(row=r, column=21).value or "").strip(),
            })

    # Flatten Counter → dict for JSON
    for short in result:
        result[short]["utm_sources"] = dict(result[short]["utm_sources"].most_common())

    pipeline_total = sum(v["enquiries_pipeline"] for v in result.values())
    week_total = sum(v["enquiries_this_week"] for v in result.values())
    print(f"[MWCC Ops]   Pipeline total: {pipeline_total} enquiries · This week: {week_total}")
    for short, v in result.items():
        print(f"[MWCC Ops]     {short}: pipeline {v['enquiries_pipeline']} · this week {v['enquiries_this_week']}")
    return dict(result)


# ─────────────────────────────────────────────────────────────────
# StartersAndExitsReport.xlsx parser — SOURCE OF TRUTH for enrolments/exits
# ─────────────────────────────────────────────────────────────────
#
# OWNA splits enrolment movements into a dedicated export with two
# sections: "New Starters" + "Exits & Upcoming Exits". Each row has a
# Date showing when the start/exit happened. We extract per-centre
# counts (both rolling + this-week) and named lists.

def _parse_starters_exits(path, period_start=None, period_end=None):
    """Parse StartersAndExitsReport.xlsx → per-centre starter + exit counts.

    Returns dict keyed by short centre name:
      {
        "Waikiki": {
          "starters_this_week":  0,
          "starters_rolling":    1,
          "exits_this_week":     1,
          "exits_rolling":       3,
          "starters_list":       [...],  # this-week only
          "exits_list":          [...],  # this-week only
        },
        ...
      }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    print(f"[MWCC Ops] Parsing StartersAndExits (SOURCE OF TRUTH for enrolments/exits): {path.name}")
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[MWCC Ops] ⚠️  Failed to open StartersAndExitsReport.xlsx: {e}")
        return {}

    ws = wb.active
    from collections import defaultdict
    result = defaultdict(lambda: {
        "starters_this_week": 0,
        "starters_rolling":   0,
        "exits_this_week":    0,
        "exits_rolling":      0,
        "starters_list":      [],
        "exits_list":         [],
    })

    # Walk rows, identify section by header
    # Section 1: "New Starters" header (col 0), centre+child+date rows follow
    # Section 2: "Exits & Upcoming Exits" header, similar structure
    current_section = None  # 'starters' or 'exits'
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not row or all(v is None or v == "" for v in row):
            continue
        first = str(row[0] or "").strip()

        # Section markers
        if first == "New Starters":
            current_section = "starters"
            continue
        if first.startswith("Exits"):
            current_section = "exits"
            continue
        if first.startswith("Total Results") or first in ("Centre", ""):
            continue
        if first.startswith("Session Difference") or first in ("Monday", "Tuesday"):
            continue
        if "(" in first and "%" in first:  # the "% change" header row
            continue

        # Data row — should be a centre name in col 0
        short = _short_centre(first)
        if not short:
            continue

        # Date is in the last meaningful column. Find it by scanning.
        date_val = None
        for v in row[::-1]:
            if v and isinstance(v, str):
                d = _parse_iso_date(v)
                if d:
                    date_val = d
                    break

        in_window = (
            period_start and period_end and date_val
            and period_start <= date_val <= period_end
        )
        record = {
            "centre": short,
            "child":  row[1] if len(row) > 1 else "",
            "date":   str(date_val) if date_val else "",
        }

        if current_section == "starters":
            result[short]["starters_rolling"] += 1
            if in_window:
                result[short]["starters_this_week"] += 1
                result[short]["starters_list"].append(record)
        elif current_section == "exits":
            result[short]["exits_rolling"] += 1
            if in_window:
                result[short]["exits_this_week"] += 1
                result[short]["exits_list"].append(record)

    # ── Deduplicate by child name (08 Jun 2026, Tia direction) ──
    # OWNA records a centre transfer as a "starter" at the destination centre
    # AND optionally a "starter" at the origin centre if mid-week. The same
    # child can appear in multiple centres' starter lists across a 5-week
    # window. For accurate unique-child counts, deduplicate by name within
    # the rolling window. Per-centre raw counts are preserved (each centre's
    # OWNA record is real); only the network total is deduplicated.
    seen_starters = set()
    seen_exits    = set()
    unique_starters_rolling = 0
    unique_exits_rolling    = 0
    unique_starters_week    = 0
    unique_exits_week       = 0
    for centre_data in result.values():
        for s in centre_data.get("starters_list", []):
            n = (s.get("child") or "").strip().lower()
            if n and n not in seen_starters:
                seen_starters.add(n)
                unique_starters_week += 1
        for e in centre_data.get("exits_list", []):
            n = (e.get("child") or "").strip().lower()
            if n and n not in seen_exits:
                seen_exits.add(n)
                unique_exits_week += 1

    # For the rolling window we need to inspect every starter/exit row
    # (not just this-week's). We don't have all rolling records on
    # result[centre]["starters_list"] (only this-week's). Re-walk the file
    # for accurate rolling unique counts.
    seen_starters_r = set()
    seen_exits_r    = set()
    current_section = None
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not row or all(v is None or v == "" for v in row):
            continue
        first = str(row[0] or "").strip()
        if first == "New Starters":   current_section = "starters"; continue
        if first.startswith("Exits"): current_section = "exits"; continue
        if first.startswith("Total Results") or first in ("Centre", "") or first.startswith("Session Difference") or first in ("Monday", "Tuesday") or ("(" in first and "%" in first):
            continue
        if not _short_centre(first):
            continue
        child = (row[1] if len(row) > 1 else "")
        child_key = str(child or "").strip().lower()
        if not child_key:
            continue
        if current_section == "starters":
            seen_starters_r.add(child_key)
        elif current_section == "exits":
            seen_exits_r.add(child_key)
    unique_starters_rolling = len(seen_starters_r)
    unique_exits_rolling    = len(seen_exits_r)

    starters_wk = sum(v["starters_this_week"] for v in result.values())
    exits_wk    = sum(v["exits_this_week"]    for v in result.values())
    starters_rl = sum(v["starters_rolling"]   for v in result.values())
    exits_rl    = sum(v["exits_rolling"]      for v in result.values())
    print(f"[MWCC Ops]   This week (raw): {starters_wk} starters · {exits_wk} exits (net {starters_wk - exits_wk:+d})")
    print(f"[MWCC Ops]   This week (unique children): {unique_starters_week} starters · {unique_exits_week} exits")
    print(f"[MWCC Ops]   Rolling (raw):   {starters_rl} starters · {exits_rl} exits")
    print(f"[MWCC Ops]   Rolling (unique children): {unique_starters_rolling} starters · {unique_exits_rolling} exits")

    # Stash unique counts as a tuple — caller pulls them off separately.
    # Returning the dict + dedup info as a tuple keeps centre iteration clean.
    return dict(result), {
        "unique_starters_week":    unique_starters_week,
        "unique_exits_week":       unique_exits_week,
        "unique_starters_rolling": unique_starters_rolling,
        "unique_exits_rolling":    unique_exits_rolling,
    }


# ─────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────

def parse_mwcc_ops(inbox_path=None):
    inbox = Path(inbox_path) if inbox_path else INBOX_DIR

    if not inbox.exists():
        msg = f"mwcc-inbox/ not found at {inbox}"
        print(f"[MWCC Ops] ⚠️  {msg}")
        print(f"[MWCC Ops]    Create it and drop MYWORLD_REPORT.xlsx + utilisation.xlsx there.")
        _write_empty(msg)
        return None

    report_file      = _find_file(inbox, "MYWORLD_REPORT")
    utilisation_file = _find_file(inbox, "utilisation")

    if not report_file and not utilisation_file:
        msg = "No xlsx files found in mwcc-inbox/"
        print(f"[MWCC Ops] ⚠️  {msg} — dropping empty ops state.")
        _write_empty(msg)
        return None

    # --- Parse MYWORLD_REPORT ---
    ops_data     = {}
    period_start = None
    period_end   = None

    if report_file:
        centres_data, period_start, period_end = _parse_myworld_report(report_file)
        if centres_data:
            ops_data = centres_data
    else:
        print("[MWCC Ops] ⚠️  MYWORLD_REPORT*.xlsx not found — revenue/wage/enquiry data missing")

    # --- Parse utilisation and merge room detail ---
    if utilisation_file:
        util_data = _parse_utilisation(utilisation_file)
        for centre_name, util in util_data.items():
            if centre_name in ops_data:
                # Merge: utilisation gives daily granularity, overrides MYWORLD_REPORT room %s
                ops_data[centre_name]["rooms_detail"]   = util["rooms"]
                ops_data[centre_name]["daily_occ_pcts"] = util["daily_occ_pcts"]
            else:
                # Centre appears only in utilisation (no financial data this week)
                ops_data[centre_name] = {
                    "name":           centre_name,
                    "rooms_detail":   util["rooms"],
                    "daily_occ_pcts": util["daily_occ_pcts"],
                }
    else:
        print("[MWCC Ops] ⚠️  utilisation*.xlsx not found — per-room daily data missing")

    # --- SOURCE OF TRUTH OVERRIDES (08 Jun 2026) ---
    # leads-enquiries.xlsx is the source of truth for enquiries (with UTM
    # attribution). StartersAndExitsReport.xlsx is the source of truth for
    # enrolments + exits (with named lists). These OVERRIDE the counts that
    # came from MYWORLD_REPORT, which were too aggregated to be reliable.
    leads_file   = _find_file(inbox, "leads-enquiries")
    starters_file = _find_file(inbox, "StartersAndExitsReport")

    # Use the OWNA Mon-Fri window for date filtering
    canonical_start_str, canonical_end_str = compute_owna_week()
    canonical_start_d = _parse_iso_date(canonical_start_str)
    canonical_end_d   = _parse_iso_date(canonical_end_str)

    if leads_file:
        leads_data = _parse_leads_enquiries(leads_file, canonical_start_d, canonical_end_d)
        for short_centre, leads in leads_data.items():
            if short_centre not in ops_data:
                ops_data[short_centre] = {"name": short_centre}
            # OVERRIDE: enquiries from leads-enquiries.xlsx (source of truth)
            ops_data[short_centre]["enquiries"]          = leads["enquiries_this_week"]
            ops_data[short_centre]["enquiries_pipeline"] = leads["enquiries_pipeline"]
            ops_data[short_centre]["enquiry_utm_sources"] = leads["utm_sources"]
            ops_data[short_centre]["enquiry_leads"]      = leads["leads"]
            ops_data[short_centre]["enquiries_source"]   = "leads-enquiries.xlsx"
    else:
        print("[MWCC Ops] ⚠️  leads-enquiries.xlsx not found — enquiry counts fall back to MYWORLD_REPORT (may be inaccurate)")

    starters_dedup = None
    if starters_file:
        starters_data, starters_dedup = _parse_starters_exits(
            starters_file, canonical_start_d, canonical_end_d)
        for short_centre, se in starters_data.items():
            if short_centre not in ops_data:
                ops_data[short_centre] = {"name": short_centre}
            # OVERRIDE: enrolments + exits from StartersAndExitsReport.xlsx (source of truth)
            ops_data[short_centre]["enrolments"]           = se["starters_this_week"]
            ops_data[short_centre]["enrolments_rolling"]   = se["starters_rolling"]
            ops_data[short_centre]["exits"]                = se["exits_this_week"]
            ops_data[short_centre]["exits_rolling"]        = se["exits_rolling"]
            ops_data[short_centre]["net_movement"]         = se["starters_this_week"] - se["exits_this_week"]
            ops_data[short_centre]["starters_list"]        = se["starters_list"]
            ops_data[short_centre]["exits_list"]           = se["exits_list"]
            ops_data[short_centre]["enrolments_source"]    = "StartersAndExitsReport.xlsx"
    else:
        print("[MWCC Ops] ⚠️  StartersAndExitsReport.xlsx not found — enrolment/exit counts fall back to MYWORLD_REPORT (may be inaccurate)")

    # --- Network summary ---
    total_enquiries  = sum(c.get("enquiries",  0) for c in ops_data.values())
    total_exits      = sum(c.get("exits",      0) for c in ops_data.values())
    total_enrolments = sum(c.get("enrolments", 0) for c in ops_data.values())
    total_revenue    = sum(c.get("revenue",    0) for c in ops_data.values())

    breach_centres = [
        n for n, c in ops_data.items() if c.get("wage_breach", False)
    ]
    # Per Kelley rule (09 Jun 2026): per-room over-100% is operational
    # rebalance, NOT compliance. compliance_risk_rooms is now always empty
    # at the per-room level. True centre-level compliance check is a
    # future addition pending licensed_centre_capacity from Tia.
    compliance_risk_rooms: list[str] = []  # legal compliance — centre-level only (TODO)
    rooms_needing_rebalance = [
        f"{centre} — {room}"
        for centre, c in ops_data.items()
        for room, d in c.get("rooms_detail", {}).items()
        if d.get("staffing_rebalance_needed", False)
    ]

    # Override the period from the OWNA file with the canonical Mon-Fri window.
    # Convention: centres operate Mon-Fri only. OWNA's native Mon-Sun export
    # always has zeros for Sat/Sun anyway, so trimming to Mon-Fri removes noise
    # and aligns with how the team reports the business. Computed deterministically
    # from today's date so the dashboard always shows the canonical "last completed
    # working week" regardless of what the OWNA export's header row contains.
    file_period_start, file_period_end = period_start, period_end
    period_start, period_end = compute_owna_week()
    period_label = f"{period_start} to {period_end}"

    output = {
        "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "brand":        "mwcc",
        "period": {
            "start":  period_start,
            "end":    period_end,
            "label":  period_label,
            "note":   "OWNA centre operations data is Mon–Fri (centres operate weekdays only). Digital marketing scripts (GA4, GSC, Meta, Google Ads, Metricool) report Sat–Fri to capture weekend digital activity.",
            "owna_file_period": f"{file_period_start} to {file_period_end}" if file_period_start else "unknown",
        },
        "network_summary": {
            "total_enquiries":            total_enquiries,
            "total_enrolments":           total_enrolments,
            "total_exits":                total_exits,
            "net_movement":               total_enrolments - total_exits,
            "total_revenue":              round(total_revenue, 2),
            "centres_in_wage_breach":     breach_centres,
            "rooms_with_compliance_risk": compliance_risk_rooms,    # always [] now; centre-level check TODO
            "rooms_needing_rebalance":    rooms_needing_rebalance,  # operational signal for Kelley
            # Unique-child counts (deduplicated across centres for transfers)
            # Per Tia direction 08 Jun 2026 — when a child transfers between
            # centres OWNA records them twice. Counting unique names gives
            # the true number of children moving in/out of the network.
            **(starters_dedup or {}),
        },
        "centres": ops_data,
        "source_files": {
            "myworld_report": report_file.name      if report_file      else None,
            "utilisation":    utilisation_file.name if utilisation_file else None,
        },
    }

    STATE_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(output, indent=2))

    # --- Console summary ---
    print(f"\n[MWCC Ops] ✅ Saved → {OUTPUT_FILE}")
    print(f"[MWCC Ops] Period      : {period_label}")
    print(f"[MWCC Ops] Enquiries   : {total_enquiries}  |  Enrolments: {total_enrolments}  |  Exits: {total_exits}  |  Net: {total_enrolments - total_exits:+d}")
    print(f"[MWCC Ops] Revenue     : ${total_revenue:,.2f}")
    if breach_centres:
        print(f"[MWCC Ops] ⚠️  Wage breach (>{WAGE_BREACH_THRESHOLD}%): {', '.join(breach_centres)}")
    if rooms_needing_rebalance:
        print(f"[MWCC Ops] ↻ Rooms >100% (Kelley to rebalance, NOT compliance): {', '.join(rooms_needing_rebalance)}")

    return output


def _write_empty(reason: str):
    output = {
        "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "brand":        "mwcc",
        "available":    False,
        "skip_reason":  reason,
    }
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(output, indent=2))


if __name__ == "__main__":
    inbox_arg = sys.argv[1] if len(sys.argv) > 1 else None
    parse_mwcc_ops(inbox_arg)
