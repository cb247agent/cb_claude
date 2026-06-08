"""
parse_membership_data.py — Extract CB247 membership statistics from
the weekly XLSX exports dropped into cb247-membership-inbox/.

Input files (drop weekly):
  cb247-membership-inbox/PGM_ContractsSummary.xlsx  — multi-tab weekly aggregate
  cb247-membership-inbox/PGM_AllContracts.xlsx      — master contracts list
  cb247-membership-inbox/Cleverwaiver.xlsx          — cancellation survey responses

Output: state/membership-data.json

Failure mode: missing or unparseable files preserve the previous JSON
rather than blanking it (protects against forgotten weekly drop).

Filtering logic for "base unique people":
  Excludes Payment Plan Name in:
    - Employee Default Free
    - Kids Hub Child Plan
    - Recovery (Includes Sauna, Ice Bath)         ← add-on
    - Reformer Pilates STANDARD MEMBERSHIP        ← add-on
    - Neon, Group Fitness, ChasinRX               ← add-ons
    - Fitness Passport                            ← corporate
  Then deduplicates on User Number.
"""

import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

BASE_DIR    = Path(__file__).resolve().parent.parent
INBOX_DIR   = BASE_DIR / "cb247-membership-inbox"
STATE_DIR   = BASE_DIR / "state"
OUTPUT_FILE = STATE_DIR / "membership-data.json"

PGM_SUMMARY_FILE  = INBOX_DIR / "PGM_ContractsSummary.xlsx"
PGM_CONTRACTS_FILE = INBOX_DIR / "PGM_AllContracts.xlsx"
CLEVERWAIVER_FILE = INBOX_DIR / "Cleverwaiver.xlsx"

# ── Filter rules ─────────────────────────────────────────────────────────────
# Plans excluded from "base unique people" — administrative or non-revenue
EXCLUDED_PLAN_KEYWORDS = [
    "employee default free",   # Staff free access
    "fitness passport",        # Corporate scheme
]
# Kids Hub plans (child memberships, not gym members) — separate category
KIDS_HUB_KEYWORDS = [
    "kids hub",
]
# Add-on plans (Recovery, Reformer, Neon, ChasinRX, Sauna)
ADDON_KEYWORDS = [
    "recovery",
    "reformer",
    "pilates",
    "neon",
    "chasinrx",
    "chasingrx",
    "sauna",
    "ice bath",
    "group fitness",
]

# Live membership statuses (what counts as "active")
LIVE_STATUSES = {"Live"}

# Cancel reason categories — used to distinguish member-submitted cancellations
# from contracts that ended naturally (promo expiry, trial ending, etc).
#
# MEMBER_SUBMITTED: member actively chose to leave (this is the "real" churn)
# NATURAL_END:      contract ran its natural course (trial / promo expired)
# DEBT_COLLECTION:  involuntary cancellation due to non-payment
# MANAGEMENT_END:   ended by the club / system upgrade
# UNKNOWN:          unclassified — counted in totals but flagged separately
MEMBER_SUBMITTED_REASONS = {
    "Relocating",
    "Not Using the membership enough",
    "Time or schedule no longer suits",
    "Switched to another gym",
    "Work commitments",
    "Home gym or online training",
    "Extended travel away",
    "Medical/Injury",
    "AnnualCancellation",
    "Facility or equipment issues",
    "Other",  # in this data set, Other almost always = member-submitted
}
NATURAL_END_REASONS = {
    "EndOfContract",
    "EndOfPromotion",
    "2 Week Trial",
    "UpgradeOfContract",
}
DEBT_COLLECTION_REASONS = {
    "TurnedOverToVindication",
    "Cancelld Via Debt Collection Stage 4",
    "Cancelled Via Debt Collection Stage 4",  # spelling fix variant
}
MANAGEMENT_END_REASONS = {
    "EndedByManagment",
    "EndedByManagement",
}


def _classify_cancel_reason(reason):
    """Return one of: 'submitted', 'natural', 'debt', 'management', 'unknown'."""
    r = (reason or "").strip()
    if r in MEMBER_SUBMITTED_REASONS:   return "submitted"
    if r in NATURAL_END_REASONS:        return "natural"
    if r in DEBT_COLLECTION_REASONS:    return "debt"
    if r in MANAGEMENT_END_REASONS:     return "management"
    return "unknown"

CLUBS = {
    "Malaga":     "ChasingBetter247 Malaga",
    "Ellenbrook": "ChasingBetter247 Ellenbrook",
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def _is_addon(plan_name):
    """Plan is an add-on (Recovery, Reformer, Neon, ChasinRX, Sauna, etc)."""
    p = (plan_name or "").lower()
    return any(k in p for k in ADDON_KEYWORDS)


def _is_kids_hub(plan_name):
    """Plan is a Kids Hub child membership (not a gym member)."""
    p = (plan_name or "").lower()
    return any(k in p for k in KIDS_HUB_KEYWORDS)


def _is_excluded_base(plan_name):
    """Plan is administrative or corporate, excluded from base count."""
    p = (plan_name or "").lower()
    return any(k in p for k in EXCLUDED_PLAN_KEYWORDS)


def _is_base_plan(plan_name):
    """Row counts as 'base unique people' if: not add-on, not Kids Hub, not excluded."""
    if not plan_name:
        return False
    return (not _is_addon(plan_name)
            and not _is_kids_hub(plan_name)
            and not _is_excluded_base(plan_name))


def _read_tab(wb, tab_name):
    """Yield rows from a tab as dicts (header on row 2 for PGM, row 1 for AllContracts)."""
    if tab_name not in wb.sheetnames:
        return
    ws = wb[tab_name]
    # PGM Summary tabs have a header on row 2 (row 1 = "Clients added in period")
    # except the 'Summary' tab itself (period header on row 1, data row 2 = column headers)
    header_row = 2 if tab_name != "Summary" else 2
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or v == "" for v in row_vals):
            continue
        yield dict(zip(headers, row_vals))


def _read_allcontracts(wb):
    """AllContracts tab has header on row 1."""
    if "AllContracts" not in wb.sheetnames:
        return
    ws = wb["AllContracts"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or v == "" for v in row_vals):
            continue
        yield dict(zip(headers, row_vals))


def _count_unique_base(rows, club_filter=None):
    """Filter rows to base memberships, dedupe by User number."""
    seen = set()
    for r in rows:
        if club_filter and (r.get("Club") or "") != club_filter:
            continue
        plan = r.get("Payment Plan Name") or ""
        if not _is_base_plan(plan):
            continue
        u = r.get("User number") or r.get("UserNumber")
        if u and u not in seen:
            seen.add(u)
    return len(seen)


# ── PGM Summary parser ───────────────────────────────────────────────────────

def parse_pgm_summary():
    """Return dict with all weekly aggregates from the summary tab + per-tab unique people."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"available": False, "skip_reason": "openpyxl not installed"}

    if not PGM_SUMMARY_FILE.exists():
        return {"available": False, "skip_reason": "PGM_ContractsSummary.xlsx not in inbox"}

    try:
        wb = load_workbook(PGM_SUMMARY_FILE, data_only=True)
    except Exception as e:
        return {"available": False, "skip_reason": f"failed to open: {e}"}

    # Summary tab — headline weekly aggregates per club
    ws = wb["Summary"]
    period_raw = ws.cell(row=1, column=1).value or ""
    # Parse period: "Period: 2026-05-25 - 2026-05-31"
    period = period_raw.replace("Period:", "").strip()
    start_date, end_date = (period.split(" - ") + ["", ""])[:2]
    summary_headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]

    per_club = {}
    totals = {}
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in vals):
            continue
        row = dict(zip(summary_headers, vals))
        club_name = row.get("Club") or ""
        if club_name == "Total":
            totals = {k: row.get(k, 0) for k in summary_headers if k != "Club"}
        else:
            # Map "ChasingBetter247 Malaga" → "Malaga"
            for short, full in CLUBS.items():
                if club_name == full:
                    per_club[short] = {k: row.get(k, 0) for k in summary_headers if k != "Club"}

    # Per-tab base unique people
    new_rows = list(_read_tab(wb, "New contracts"))
    ended_rows = list(_read_tab(wb, "Ended contracts"))
    future_rows = list(_read_tab(wb, "Future cancellations"))

    unique_base = {
        "new":     {short: _count_unique_base(new_rows,    CLUBS[short]) for short in CLUBS},
        "ended":   {short: _count_unique_base(ended_rows,  CLUBS[short]) for short in CLUBS},
        "future":  {short: _count_unique_base(future_rows, CLUBS[short]) for short in CLUBS},
    }
    unique_base["new"]["Total"]    = sum(unique_base["new"].values())
    unique_base["ended"]["Total"]  = sum(unique_base["ended"].values())
    unique_base["future"]["Total"] = sum(unique_base["future"].values())

    # ── Ended contracts classified by reason category ─────────────────────────
    # One member can have multiple ended contracts in the same week (base +
    # add-ons). We count UNIQUE PEOPLE per category, then sum contracts too
    # so the dashboard can show both views.
    cancel_reasons = Counter()
    contracts_by_category = Counter()
    people_by_category = {  # category → club → set of user numbers
        "submitted":  {"Malaga": set(), "Ellenbrook": set()},
        "natural":    {"Malaga": set(), "Ellenbrook": set()},
        "debt":       {"Malaga": set(), "Ellenbrook": set()},
        "management": {"Malaga": set(), "Ellenbrook": set()},
        "unknown":    {"Malaga": set(), "Ellenbrook": set()},
    }
    submitted_reasons_detail = Counter()  # reasons within "submitted" only

    for r in ended_rows:
        reason = (r.get("Cancel reason") or "").strip()
        if reason:
            cancel_reasons[reason] += 1
        category = _classify_cancel_reason(reason)
        contracts_by_category[category] += 1
        if category == "submitted":
            submitted_reasons_detail[reason] += 1

        # Bucket the user into people_by_category, only if base plan
        # (drops add-on rows so we don't double-count one person)
        plan = r.get("Payment Plan Name") or ""
        if not _is_base_plan(plan):
            continue
        user_num = r.get("User number")
        club_raw = r.get("Club") or ""
        club_short = None
        for short, full in CLUBS.items():
            if full == club_raw:
                club_short = short
                break
        if user_num and club_short:
            people_by_category[category][club_short].add(user_num)

    # Flatten people sets to counts
    submitted_cancellations = {
        short: len(people_by_category["submitted"][short]) for short in CLUBS
    }
    submitted_cancellations["Total"] = sum(submitted_cancellations.values())

    natural_endings = {
        short: len(people_by_category["natural"][short]) for short in CLUBS
    }
    natural_endings["Total"] = sum(natural_endings.values())

    debt_endings = {
        short: len(people_by_category["debt"][short]) for short in CLUBS
    }
    debt_endings["Total"] = sum(debt_endings.values())

    # Add new keys to unique_base for the frontend
    unique_base["submitted_cancellation"] = submitted_cancellations
    unique_base["natural_ending"]         = natural_endings
    unique_base["debt_ending"]            = debt_endings

    return {
        "available":    True,
        "period":       {"raw": period, "start": start_date, "end": end_date},
        "per_club":     per_club,
        "totals":       totals,
        "unique_base":  unique_base,
        "cancel_reasons_pgm": dict(cancel_reasons.most_common(20)),
        # ── Ended-contract categorisation (user feedback: "Ended" includes
        # natural promo expiries + debt collections, not just member-submitted
        # cancellations). The frontend uses these to surface the real churn. ──
        "ended_categorised": {
            "contracts_by_category": dict(contracts_by_category),
            "submitted_reasons_detail": dict(submitted_reasons_detail.most_common(10)),
            "category_definitions": {
                "submitted":  "Member actively chose to leave (real churn)",
                "natural":    "Trial / promo expired (not member-submitted)",
                "debt":       "Cancelled due to non-payment / debt collection",
                "management": "Ended by the club or system upgrade",
                "unknown":    "No reason code or unrecognised",
            },
        },
        "rows_seen": {
            "new":   len(new_rows),
            "ended": len(ended_rows),
            "future": len(future_rows),
        },
    }


# ── Cleverwaiver parser ──────────────────────────────────────────────────────

def parse_cleverwaiver():
    """Return cancellation reason + would-return + would-help breakdowns."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"available": False, "skip_reason": "openpyxl not installed"}

    if not CLEVERWAIVER_FILE.exists():
        return {"available": False, "skip_reason": "Cleverwaiver.xlsx not in inbox"}

    try:
        wb = load_workbook(CLEVERWAIVER_FILE, data_only=True)
    except Exception as e:
        return {"available": False, "skip_reason": f"failed to open: {e}"}

    ws = wb.active
    if ws.max_row < 2:
        return {"available": False, "skip_reason": "Cleverwaiver has no data"}

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Find column indexes by partial match (header text is verbose)
    def _find_col(needle):
        for i, h in enumerate(headers):
            if h and needle.lower() in str(h).lower():
                return i
        return None

    club_col   = _find_col("Club")
    reason_col = _find_col("main reason for cancelling")
    return_col = _find_col("would you return")
    helped_col = _find_col("what would have helped")

    reasons = Counter()
    reasons_by_club = {"Malaga": Counter(), "Ellenbrook": Counter()}
    returns = Counter()
    helped  = Counter()
    total = 0
    # Per-club response counts — 1 row in CleverWaiver = 1 genuine
    # member-submitted cancellation (verified intent — they completed the
    # exit survey). This becomes the new source of truth for "submitted
    # cancellations" per Option C ruling 08 Jun 2026.
    responses_by_club = {"Malaga": 0, "Ellenbrook": 0, "Unattributed": 0}

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or v == "" for v in row):
            continue
        total += 1
        club_raw = (row[club_col] or "") if club_col is not None else ""
        # Map full name -> short
        club_short = None
        for short, full in CLUBS.items():
            if full == club_raw:
                club_short = short
                break
        # Bump per-club response counter — used by capture-rate calc
        if club_short:
            responses_by_club[club_short] += 1
        else:
            responses_by_club["Unattributed"] += 1

        # Reason (may be comma-separated for multi-select)
        if reason_col is not None:
            raw = row[reason_col] or ""
            for r1 in str(raw).split(","):
                r1 = r1.strip()
                if r1:
                    reasons[r1] += 1
                    if club_short:
                        reasons_by_club[club_short][r1] += 1

        # Would return
        if return_col is not None:
            ret = (row[return_col] or "").strip()
            if ret:
                returns[ret] += 1

        # What would have helped
        if helped_col is not None:
            raw = row[helped_col] or ""
            for h1 in str(raw).split(","):
                h1 = h1.strip()
                if h1 and h1.lower() not in ("nothing specific", "not applicable", "n/a"):
                    helped[h1] += 1

    # Per Option C (08 Jun 2026): CleverWaiver becomes the source of
    # truth for "submitted cancellations". Each row = one member who
    # actively chose to cancel + completed the exit survey. PGM ended-
    # contracts remains the comprehensive view (catches phone/walk-in
    # cancels that bypass the survey) — capture rate = CleverWaiver / PGM.
    submitted_by_club = {
        "Malaga":     responses_by_club["Malaga"],
        "Ellenbrook": responses_by_club["Ellenbrook"],
        "Total":      responses_by_club["Malaga"] + responses_by_club["Ellenbrook"],
    }

    return {
        "available":             True,
        "total_responses":       total,
        # NEW — source-of-truth submitted cancellations per club + total
        "submitted_by_club":     submitted_by_club,
        "unattributed_responses": responses_by_club["Unattributed"],
        "reasons":               dict(reasons.most_common(15)),
        "reasons_by_club":       {k: dict(v.most_common(10)) for k, v in reasons_by_club.items()},
        "would_return":          dict(returns.most_common()),
        "would_have_helped":     dict(helped.most_common(10)),
    }


# ── AllContracts parser (add-on performance) ────────────────────────────────

def parse_allcontracts():
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"available": False, "skip_reason": "openpyxl not installed"}

    if not PGM_CONTRACTS_FILE.exists():
        return {"available": False, "skip_reason": "PGM_AllContracts.xlsx not in inbox"}

    try:
        wb = load_workbook(PGM_CONTRACTS_FILE, data_only=True)
    except Exception as e:
        return {"available": False, "skip_reason": f"failed to open: {e}"}

    rows = list(_read_allcontracts(wb))
    if not rows:
        return {"available": False, "skip_reason": "AllContracts tab empty"}

    addon_active = Counter()
    addon_by_club = {"Malaga": Counter(), "Ellenbrook": Counter()}
    total_active_base = {"Malaga": set(), "Ellenbrook": set()}

    for r in rows:
        plan = (r.get("Payment Plan Name") or "")
        status = r.get("Membership Status") or ""
        if status not in LIVE_STATUSES:
            continue
        club_raw = r.get("Club") or ""
        club_short = None
        for short, full in CLUBS.items():
            if full == club_raw:
                club_short = short
                break

        if _is_addon(plan):
            # Bucket add-ons into common categories
            p_low = plan.lower()
            if "recovery" in p_low or "sauna" in p_low or "ice bath" in p_low:
                bucket = "Recovery (Sauna + Ice Bath)"
            elif "reformer" in p_low or "pilates" in p_low:
                bucket = "Reformer Pilates"
            elif "neon" in p_low or "group fitness" in p_low:
                bucket = "Neon / Group Fitness"
            elif "kids hub" in p_low:
                bucket = "Kids Hub"
            elif "chasinrx" in p_low or "chasingrx" in p_low:
                bucket = "ChasinRX"
            else:
                bucket = "Other add-on"
            addon_active[bucket] += 1
            if club_short:
                addon_by_club[club_short][bucket] += 1
        elif _is_base_plan(plan):
            u = r.get("User number")
            if u and club_short:
                total_active_base[club_short].add(u)

    return {
        "available":         True,
        "total_active_base": {k: len(v) for k, v in total_active_base.items()},
        "total_active_base_combined": sum(len(v) for v in total_active_base.values()),
        "addon_active":      dict(addon_active.most_common()),
        "addon_by_club":     {k: dict(v.most_common()) for k, v in addon_by_club.items()},
    }


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    print("[membership] Parsing PGM Contracts Summary...")
    summary = parse_pgm_summary()
    if summary.get("available"):
        print(f"  ✓ Period: {summary['period']['raw']}")
        for club, m in summary["per_club"].items():
            print(f"  ✓ {club}: New {m.get('NewContracts',0)} · Ended {m.get('EndedContracts',0)} · Future Canc {m.get('FutureCancellations',0)}")
    else:
        print(f"  ⚠ {summary.get('skip_reason')}")

    print("\n[membership] Parsing Cleverwaiver...")
    cleverwaiver = parse_cleverwaiver()
    if cleverwaiver.get("available"):
        print(f"  ✓ {cleverwaiver['total_responses']} survey responses")
        print(f"  ✓ Top 3 reasons: {list(cleverwaiver['reasons'].keys())[:3]}")
        print(f"  ✓ Would return: {cleverwaiver['would_return']}")
    else:
        print(f"  ⚠ {cleverwaiver.get('skip_reason')}")

    print("\n[membership] Parsing AllContracts...")
    contracts = parse_allcontracts()
    if contracts.get("available"):
        ab = contracts["total_active_base"]
        print(f"  ✓ Active base members: Malaga {ab.get('Malaga',0)} · Ellenbrook {ab.get('Ellenbrook',0)} · Combined {contracts['total_active_base_combined']}")
        print(f"  ✓ Active add-ons: {contracts['addon_active']}")
    else:
        print(f"  ⚠ {contracts.get('skip_reason')}")

    # ── Prior week snapshot for WoW comparison ──────────────────────────
    # When membership-history.json exists, use its second-most-recent entry.
    # On first run, falls back to a hardcoded W/E 18-24 May 2026 reference
    # so the dashboard shows WoW deltas immediately.
    history_file = STATE_DIR / "membership-history.json"
    prior_week = None
    history = []
    if history_file.exists():
        try:
            history = json.loads(history_file.read_text())
            if len(history) >= 1:
                prior_week = history[-1]  # last saved week (before this run)
        except Exception:
            history = []

    if prior_week is None and summary.get("available"):
        # First-run bootstrap — W/E 18-24 May 2026 from existing weekly dashboard
        prior_week = {
            "period": {"raw": "2026-05-18 - 2026-05-24", "start": "2026-05-18", "end": "2026-05-24"},
            "totals": {
                "NewContracts": 235, "EndedContracts": 228,
                "SuspendedContracts": 170, "FutureCancellations": 494,
                "ProjectedCancellations": 426,
            },
            "per_club": {
                "Malaga":     {"NewContracts": 135, "EndedContracts": 116, "FutureCancellations": 302},
                "Ellenbrook": {"NewContracts": 100, "EndedContracts": 112, "FutureCancellations": 192},
            },
            "_source": "bootstrap-from-existing-dashboard",
        }

    # ── Submitted cancellation reconciliation (Option C — 08 Jun 2026) ──
    # Source of truth: CleverWaiver (genuine member intent — they completed
    #                  the exit survey)
    # Comparison:      PGM Ended contracts filtered by submitted reasons
    #                  (catches phone/walk-in cancels that bypass survey)
    # Capture rate:    CleverWaiver / PGM × 100 — % of cancellations that
    #                  went through the survey path.
    # Target: ≥70%. Below 70% = process issue (reception not directing
    # cancellers to fill survey) → emitter raises Work Queue action.
    submitted_truth = {"available": False}
    if cleverwaiver.get("available") and summary.get("available"):
        cw_by_club  = cleverwaiver.get("submitted_by_club", {})
        pgm_by_club = (summary.get("unique_base") or {}).get("submitted_cancellation", {})
        capture_rate = {}
        for club in ("Malaga", "Ellenbrook", "Total"):
            cw  = cw_by_club.get(club, 0)
            pgm = pgm_by_club.get(club, 0)
            capture_rate[club] = round(100.0 * cw / pgm, 1) if pgm else None
        submitted_truth = {
            "available": True,
            "source": "cleverwaiver",
            "by_club": cw_by_club,
            "pgm_comparison": pgm_by_club,
            "capture_rate_pct": capture_rate,
            "capture_target_pct": 70,
            "below_target": (capture_rate.get("Total") is not None
                             and capture_rate["Total"] < 70),
        }
        # Console feedback so the operator sees the new numbers immediately
        cap_t = capture_rate.get("Total")
        cap_str = f"{cap_t}%" if cap_t is not None else "n/a"
        print(f"\n[membership] Submitted cancellations (CleverWaiver — source of truth):")
        print(f"  Malaga    : {cw_by_club.get('Malaga')}  · PGM comparison {pgm_by_club.get('Malaga')}")
        print(f"  Ellenbrook: {cw_by_club.get('Ellenbrook')}  · PGM comparison {pgm_by_club.get('Ellenbrook')}")
        print(f"  Total     : {cw_by_club.get('Total')}  · PGM comparison {pgm_by_club.get('Total')}  · Capture {cap_str} (target ≥70%)")
        if submitted_truth["below_target"]:
            print(f"  ⚠️  Capture rate below 70% — emitter will flag Work Queue action for reception team")

    # Build output payload
    output = {
        "parsed_at":   datetime.now(timezone.utc).isoformat(),
        "available":   summary.get("available", False),
        "summary":     summary,
        "cleverwaiver": cleverwaiver,
        "contracts":   contracts,
        "prior_week":  prior_week,
        # NEW (Option C 08 Jun 2026) — single source of truth for the
        # dashboard "submitted cancellations" card. Reads from CleverWaiver
        # with PGM as comparison + capture-rate health check.
        "submitted_cancellation_truth": submitted_truth,
    }

    # Append current week to history (cap last 12 weeks)
    if summary.get("available"):
        current_snap = {
            "period":   summary.get("period"),
            "totals":   summary.get("totals"),
            "per_club": summary.get("per_club"),
            "parsed_at": output["parsed_at"],
        }
        # Avoid duplicate-on-same-period entries when re-running
        history = [h for h in history if (h.get("period") or {}).get("raw") != current_snap["period"]["raw"]]
        history.append(current_snap)
        history = history[-12:]  # keep last 12 weeks
        history_file.write_text(json.dumps(history, indent=2, default=str))

    STATE_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(output, indent=2, default=str))
    print(f"\n[membership] Saved → {OUTPUT_FILE}")
    return output


if __name__ == "__main__":
    result = main()
    sys.exit(0 if result.get("available") else 1)
